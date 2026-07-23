#!/usr/bin/env python3
"""
gbp_pitch_list.py — build the follow-up-complaint pitch list (review-scan step 4, final).

Reads the classifiers' flag files, joins each flagged business back to the store (audit + contact
data), drops the big nationals (review ceiling), ranks so the highest-rated shops with a real leak
lead, and writes an XLSX with one tab per trade (excavation / septic / plumbers, closest fit first)
plus an "All" tab, and a flat CSV.

Ranking rule (2026-07-05): rating DESC, then follow-up-complaint count DESC, then fewest total
reviews (owner-operated) — "highest-rated shops with a real-but-small leak first", not raw volume.

  python tools/gbp_pitch_list.py                                  # defaults: cap 500 reviews, >=2 complaints
  python tools/gbp_pitch_list.py --max-reviews 0 --min-complaints 1   # keep everything
"""

from __future__ import annotations

import argparse
import csv
import glob
import json
import sqlite3
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import gbp_trades

DB_PATH = Path("data/leads.sqlite")
COLUMNS = [
    "Rank",
    "Business Name",
    "Trade",
    "Category",
    "State",
    "Rating",
    "Reviews",
    "Neg reviews",
    "Followup complaints",
    "Complaint types",
    "Evidence 1",
    "Evidence 2",
    "Phone",
    "Email",
    "Owner",
    "Website",
    "Google Maps",
]


def load_flags(flag_dir: Path) -> dict:
    flagged = {}
    for f in sorted(glob.glob(str(flag_dir / "batch_*.json"))):
        raw = Path(f).read_text(encoding="utf-8").strip()
        lo, hi = raw.find("["), raw.rfind("]")  # tolerate a stray preamble/fence
        try:
            arr = json.loads(raw[lo : hi + 1]) if lo != -1 and hi != -1 else json.loads(raw)
        except ValueError as e:
            print(f"  ! {Path(f).name}: {e}", file=sys.stderr)
            continue
        for b in arr:
            if b.get("place_id"):
                flagged[b["place_id"]] = b
    return flagged


def build_rows(conn, flagged: dict, max_reviews: int, min_complaints: int) -> list[dict]:
    rows = []
    for pid, b in flagged.items():
        r = conn.execute(
            "SELECT b.name, b.phone, b.website, b.state_code, b.maps_url, g.category, "
            "g.rating_value, g.rating_votes, g.neg_reviews, ct.email, ct.owner_name "
            "FROM businesses b JOIN gbp_audits g ON g.place_id=b.place_id "
            "LEFT JOIN site_contacts ct ON ct.place_id=b.place_id "
            "WHERE b.place_id=? AND g.status='complete' ORDER BY g.id DESC LIMIT 1",
            (pid,),
        ).fetchone()
        if not r:
            continue
        cc = b.get("complaint_count") or 1
        votes = r["rating_votes"] or 0
        if cc < min_complaints:
            continue
        if max_reviews and votes > max_reviews:  # drop the big nationals
            continue
        group = gbp_trades.trade_group(r["category"])
        ev = b.get("evidence") or []
        rows.append(
            {
                "place_id": pid,
                "Business Name": r["name"],
                "Trade": group,
                "Category": r["category"],
                "State": r["state_code"],
                "Rating": r["rating_value"],
                "Reviews": votes,
                "Neg reviews": r["neg_reviews"],
                "Followup complaints": cc,
                "Complaint types": ", ".join(b.get("complaint_types") or []),
                "Evidence 1": ev[0] if len(ev) > 0 else "",
                "Evidence 2": ev[1] if len(ev) > 1 else "",
                "Phone": r["phone"] or "",
                "Email": r["email"] or "",
                "Owner": r["owner_name"] or "",
                "Website": r["website"] or "",
                "Google Maps": r["maps_url"] or "",
                "_group_rank": gbp_trades.GROUP_ORDER.index(group)
                if group in gbp_trades.GROUP_ORDER
                else 9,
            }
        )
    # rating leads; then a real leak; then owner-operated (fewest reviews).
    rows.sort(
        key=lambda x: (-(x["Rating"] or 0), -(x["Followup complaints"] or 0), x["Reviews"] or 0)
    )
    return rows


def _write_sheet(ws, rows, header_fill, Font, PatternFill, get_column_letter):
    ws.append(COLUMNS)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
    maps_i = COLUMNS.index("Google Maps")
    for i, x in enumerate(rows, 1):
        x = {**x, "Rank": i}
        row = [x.get(c, "") for c in COLUMNS]
        if x.get("Google Maps"):
            row[maps_i] = f'=HYPERLINK("{x["Google Maps"]}","map")'
        ws.append(row)
    widths = {
        "Business Name": 30,
        "Category": 20,
        "Complaint types": 30,
        "Evidence 1": 58,
        "Evidence 2": 58,
        "Phone": 16,
        "Email": 26,
        "Owner": 18,
        "Website": 32,
    }
    for name, wd in widths.items():
        ws.column_dimensions[get_column_letter(COLUMNS.index(name) + 1)].width = wd
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"


def write_xlsx(rows, out_xlsx: Path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    fill = PatternFill("solid", fgColor="15191D")
    wb = openpyxl.Workbook()
    _write_sheet(wb.active, rows, fill, Font, PatternFill, get_column_letter)
    wb.active.title = f"All ({len(rows)})"
    for group in gbp_trades.GROUP_ORDER:
        sub = [x for x in rows if x["Trade"] == group]
        if sub:
            _write_sheet(
                wb.create_sheet(f"{group.capitalize()} ({len(sub)})"),
                sub,
                fill,
                Font,
                PatternFill,
                get_column_letter,
            )
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


def write_csv(rows, out_csv: Path):
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with open(out_csv, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(COLUMNS)
        for i, x in enumerate(rows, 1):
            w.writerow([({**x, "Rank": i}).get(c, "") for c in COLUMNS])


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Build the ranked, trade-tabbed follow-up-complaint pitch list."
    )
    ap.add_argument("--flag-dir", default=".tmp/gbp/flags")
    ap.add_argument(
        "--max-reviews",
        type=int,
        default=500,
        help="Drop businesses with more than this many reviews (nationals). 0 = no cap.",
    )
    ap.add_argument(
        "--min-complaints",
        type=int,
        default=2,
        help="Require at least this many follow-up complaints (a real pattern).",
    )
    ap.add_argument("--out", default="output/excavation-followup-pitch-list.xlsx")
    ap.add_argument("--db", default=str(DB_PATH))
    args = ap.parse_args(argv)

    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row
    flagged = load_flags(Path(args.flag_dir))
    rows = build_rows(conn, flagged, args.max_reviews, args.min_complaints)
    if not rows:
        print("No businesses after filters.", file=sys.stderr)
        return 1

    out_xlsx = Path(args.out)
    write_csv(rows, out_xlsx.with_suffix(".csv"))
    try:
        write_xlsx(rows, out_xlsx)
        xlsx_note = str(out_xlsx)
    except ImportError:
        xlsx_note = "(openpyxl missing; CSV only)"

    from collections import Counter

    by_trade = Counter(x["Trade"] for x in rows)
    print(
        f"PITCH LIST: {len(rows)} businesses "
        f"(cap {args.max_reviews or '-'} reviews, >= {args.min_complaints} complaints)"
    )
    print("by trade:", {g: by_trade.get(g, 0) for g in gbp_trades.GROUP_ORDER})
    print(f"-> {xlsx_note}\n-> {out_xlsx.with_suffix('.csv')}")
    print("Top 10 (rating, then leak, then owner-operated):")
    for i, x in enumerate(rows[:10], 1):
        print(
            f"  {i:>2}. {x['Business Name'][:30]:30} {x['State']:2} {x['Trade']:10} "
            f"{x['Rating']}* {x['Reviews']:>4}rev n={x['Followup complaints']}"
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
