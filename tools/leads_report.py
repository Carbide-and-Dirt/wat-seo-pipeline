#!/usr/bin/env python3
"""
leads_report.py - segmented prospect report from the master store (HLD FR-14).

Delivers the national prospect workbook: a Summary tab, one ranked tab per
state/province, and a flat master CSV of every lead. Two ranking LENSES:

  --lens opportunity (default): the project's need-first order, applied within each
    state - CONFIRMED (match) before REVIEW (maybe); within a tier, no-website first,
    then dead/absent sites, then live sites by readiness (weaker first), WAF-blocked
    last; review_count breaks ties. (HLD FR-14.)

  --lens budget: a BUDGET-QUALIFIED order for "who can actually pay us". Tiers:
    1. PROVEN BUDGET   - already spending on marketing (running Google Ads, likely
       agency-managed, or carrying a CallRail/HubSpot/Marketo stack). Sorted
       weakest-site-first so "paying for a leaky bucket" displacement leads rise.
    2. QUALIFIED DEMAND - no paid signal, but busy enough to afford it (>= --min-reviews
       reviews) AND has a real gap (no/dead/weak site). Sorted busiest first.
    3. long tail       - everyone else (tiny/low-review or solid-site-no-spend).
    Adds 'Qualified' + 'Budget signal' columns; --qualified-only drops tier 3.
    NOTE: paid signals come from enrichment, so an un-enriched lead can only reach
    tier 2 via no-website + reviews; run enrich_sites.py first for full coverage.

Reuses the readiness/agency signals already cached in site_enrichment - computes
nothing new and spends nothing. Regenerate any time the store changes.

Usage:
    python tools/leads_report.py                                          # opportunity lens
    python tools/leads_report.py --lens budget --min-reviews 20           # budget-qualified
    python tools/leads_report.py --lens budget --qualified-only --out output/budget.xlsx
"""

import argparse
import csv
import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TOOLS = Path(__file__).resolve().parent
sys.path.insert(0, str(TOOLS))

import leads_db

from lib.common import utf8_stdout

utf8_stdout()

DEFAULT_CONFIG = "targets/excavating-national.json"

# relevance value -> human tier label + its rank (lower sorts first).
_TIER = {"match": ("CONFIRMED", 0), "maybe": ("REVIEW", 1)}
# has-website site_status values that mean "effectively no usable site" (high opportunity).
_BROKEN = {"dead", "unreachable", "social_only", "directory", "parked"}
# Marketing-stack tags that signal a business actively spends on marketing (a "can pay"
# signal). Google Ads is its own column; mirrors site_fingerprint.MANAGED_TAGS.
_MANAGED_TAGS = {"CallRail", "HubSpot", "Marketo"}
# A live readiness score at/above this = a site weak enough to count as a real "gap".
_WEAK_SITE = 5
DEFAULT_MIN_REVIEWS = 20

# Per-lead columns shared by the state tabs and the CSV. "Google Maps" holds the raw
# maps URL in a record; the xlsx writer wraps it in a HYPERLINK, the CSV keeps it raw.
LEAD_COLUMNS = [
    "Tier",
    "Business Name",
    "Reviews",
    "Rating",
    "Has Website",
    "Site status",
    "Readiness",
    "Management",
    "Conf",
    "Builder",
    "Marketing tags",
    "Agency credit",
    "Top gap",
    "Phone",
    "Email",
    "Owner",
    "Website",
    "Google Maps",
]
# Budget lens leads with the qualification + the paid signal, then the why-now + contact columns.
BUDGET_COLUMNS = [
    "Qualified",
    "Budget signal",
    "Tier",
    "Business Name",
    "Reviews",
    "Rating",
    "Has Website",
    "Site status",
    "Readiness",
    "Management",
    "Builder",
    "Marketing tags",
    "Top gap",
    "Phone",
    "Email",
    "Owner",
    "Website",
    "Google Maps",
]
# GBP-neglect lens: leads ranked by how neglected their Google Business Profile is (worst first
# = best pitch), from the gbp_audits store. Neglect + the fired signals lead; contact columns follow.
GBP_COLUMNS = [
    "Tier",
    "Business Name",
    "Neglect",
    "Claimed",
    "Top gaps",
    "GBP reviews",
    "Neg reviews",
    "Photos",
    "Days since post",
    "Rating",
    "Phone",
    "Email",
    "Owner",
    "Website",
    "Google Maps",
]
_WIDTHS = {
    "Business Name": 32,
    "Site status": 13,
    "Management": 30,
    "Builder": 16,
    "Marketing tags": 26,
    "Agency credit": 20,
    "Top gap": 46,
    "Phone": 16,
    "Email": 28,
    "Owner": 20,
    "Website": 38,
    "Google Maps": 16,
    "Qualified": 17,
    "Budget signal": 30,
    "Neglect": 9,
    "Claimed": 9,
    "Top gaps": 34,
    "Days since post": 16,
    "GBP reviews": 12,
    "Neg reviews": 12,
    "Photos": 9,
}

# Neglect-signal keys (gbp_audit.NEGLECT_WEIGHTS) -> short human labels for the report.
_GBP_SIGNAL_LABELS = {
    "unclaimed": "unclaimed",
    "stale_posts": "stale posts",
    "thin_reviews": "few reviews",
    "few_photos": "few photos",
    "no_secondary_categories": "no 2nd categories",
    "no_hours": "no hours",
    "sparse_attributes": "few attributes",
    "no_description": "no description",
}


def _opp_bucket(no_website, site_status):
    """Lower = higher up the list = clearer sales opening."""
    if no_website:
        return 0
    if site_status in _BROKEN:
        return 1
    if site_status == "live":
        return 2
    return 4  # blocked / unknown - a working site probably exists, lowest priority


def _readiness_cell(no_website, site_status, score):
    if no_website:
        return "NO SITE"
    if site_status == "live":
        return score if score is not None else "n/a"
    return f"n/a ({site_status})" if site_status else "n/a"


def _top_gap(no_website, gaps_json):
    if no_website:
        return "no website - full build opportunity"
    try:
        gaps = json.loads(gaps_json) if gaps_json else []
    except (ValueError, TypeError):
        gaps = []
    return gaps[0] if gaps else "solid site - few gaps"


def _tags(tags_json):
    try:
        return ", ".join(json.loads(tags_json)) if tags_json else ""
    except (ValueError, TypeError):
        return ""


def _tag_set(tags_json):
    try:
        return set(json.loads(tags_json)) if tags_json else set()
    except (ValueError, TypeError):
        return set()


def _sort_key(r):
    tier_rank = _TIER.get(r["relevance"], ("?", 9))[1]
    bucket = _opp_bucket(r["no_website"], r["site_status"])
    # Within live sites, weaker (higher readiness) first -> negate the score.
    score = r["readiness_score"] if r["readiness_score"] is not None else -1
    return (tier_rank, bucket, -score, -(r["review_count"] or 0), r["name"] or "")


# ---- budget-qualified lens ----


def paid_signal(r):
    """The strongest 'this business already pays for marketing' signal, or '' if none.
    Ordered by how clearly it proves spend: live ads > an agency > a marketing stack."""
    if r["google_ads"]:
        return "Google Ads"
    if r["mgmt_status"] == "likely agency-managed":
        return "agency-managed"
    managed = _tag_set(r["marketing_tags_json"]) & _MANAGED_TAGS
    if managed:
        return "/".join(sorted(managed))
    return ""


def _has_gap(r):
    """True if the lead has a real web gap worth selling against."""
    if r["no_website"]:
        return True
    if r["site_status"] and r["site_status"] != "live":
        return True
    rs = r["readiness_score"]
    return rs is not None and rs >= _WEAK_SITE


def budget_assess(r, min_reviews=DEFAULT_MIN_REVIEWS):
    """Classify a lead by ability/willingness to pay. Returns (tier, label, signal):
    1 PROVEN BUDGET (already spending), 2 QUALIFIED DEMAND (busy + has a gap),
    3 long tail (neither). 'signal' is the short why-they-qualify string."""
    sig = paid_signal(r)
    if sig:
        return 1, "PROVEN BUDGET", sig
    reviews = r["review_count"] or 0
    if reviews >= min_reviews and _has_gap(r):
        gap = (
            "no site"
            if r["no_website"]
            else (
                "dead/blocked site"
                if (r["site_status"] and r["site_status"] != "live")
                else "weak site"
            )
        )
        return 2, "QUALIFIED DEMAND", f"busy: {reviews} reviews, {gap}"
    return 3, "long tail", ""


def _budget_sort_key(r, min_reviews=DEFAULT_MIN_REVIEWS):
    tier, _, _ = budget_assess(r, min_reviews)
    rs = r["readiness_score"] if r["readiness_score"] is not None else -1
    if tier == 1:  # proven budget: weakest site first (best displacement wedge)
        return (1, -rs, -(r["review_count"] or 0), r["name"] or "")
    return (tier, -(r["review_count"] or 0), -rs, r["name"] or "")  # else busiest first


def fetch_rows(conn, state_codes=None):
    """Every lead (LEFT JOIN enrichment, so no-website leads come through) as plain
    dicts, globally sorted into the opportunity ranking. Because the sort key has no
    state term, grouping the sorted list by state preserves each state's ranking."""
    where, params = [], []
    if state_codes:
        where.append("LOWER(b.state_code) IN (%s)" % ",".join("?" for _ in state_codes))
        params.extend(state_codes)
    sql = (
        "SELECT b.place_id, b.name, b.phone, b.website, b.no_website, b.rating, "
        "b.review_count, b.relevance, b.state_code, b.state_name, b.maps_url, "
        "e.site_status, e.readiness_score, e.mgmt_status, e.mgmt_confidence, e.builder, "
        "e.marketing_tags_json, e.agency_credit, e.google_ads, e.seo_gaps_json, "
        "ct.email, ct.owner_name "
        "FROM businesses b LEFT JOIN site_enrichment e ON e.place_id = b.place_id "
        "LEFT JOIN site_contacts ct ON ct.place_id = b.place_id"
    )
    if where:
        sql += " WHERE " + " AND ".join(where)
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    rows.sort(key=_sort_key)
    return rows


# ---- GBP-neglect lens ----


def gbp_signal_label(signals_json, top=2):
    """The strongest neglect signals as a short label, e.g. 'unclaimed, few photos'."""
    try:
        fired = json.loads(signals_json) if signals_json else {}
    except (ValueError, TypeError):
        fired = {}
    on = [_GBP_SIGNAL_LABELS.get(k, k) for k, v in fired.items() if v]
    # Keep the weight order (unclaimed first) by sorting on the label map's insertion order.
    order = list(_GBP_SIGNAL_LABELS.values())
    on.sort(key=lambda x: order.index(x) if x in order else 99)
    return ", ".join(on[:top])


def fetch_gbp_rows(conn, state_codes=None):
    """Every business with a completed GBP audit, joined to its LATEST audit snapshot,
    ranked worst-profile-first (highest neglect = best pitch). Un-audited leads are
    excluded (run gbp_audit.py to populate). Sort key has no state term, so grouping the
    sorted list by state preserves each state's ranking (same contract as fetch_rows)."""
    where, params = [], []
    if state_codes:
        where.append("LOWER(b.state_code) IN (%s)" % ",".join("?" for _ in state_codes))
        params.extend(state_codes)
    sql = (
        "SELECT b.place_id, b.name, b.phone, b.website, b.no_website, b.rating, "
        "b.review_count, b.relevance, b.state_code, b.state_name, b.maps_url, "
        "g.is_claimed, g.neglect_score, g.signals_json, g.total_photos, g.rating_votes, "
        "g.neg_reviews, g.days_since_post, g.post_count, ct.email, ct.owner_name "
        "FROM businesses b "
        "JOIN (SELECT gg.* FROM gbp_audits gg "
        "      JOIN (SELECT place_id, MAX(id) mid FROM gbp_audits "
        "            WHERE status='complete' GROUP BY place_id) x "
        "      ON gg.id = x.mid) g "
        "  ON g.place_id = b.place_id "
        "LEFT JOIN site_contacts ct ON ct.place_id = b.place_id"
    )
    if where:
        sql += " WHERE " + " AND ".join(where)
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    # tier (CONFIRMED before REVIEW), then most-neglected, then busiest.
    rows.sort(
        key=lambda r: (
            _TIER.get(r["relevance"], ("?", 9))[1],
            -(r["neglect_score"] or 0),
            -(r["review_count"] or 0),
            r["name"] or "",
        )
    )
    return rows


def to_gbp_record(r, min_reviews=DEFAULT_MIN_REVIEWS):
    """One GBP-audited lead as a display record (shared by xlsx + CSV)."""
    claimed = "?" if r["is_claimed"] is None else ("yes" if r["is_claimed"] else "NO")
    dsp = r["days_since_post"]
    return {
        "Tier": _TIER.get(r["relevance"], ("?", 9))[0],
        "Business Name": r["name"],
        "Neglect": round(r["neglect_score"]) if r["neglect_score"] is not None else "",
        "Claimed": claimed,
        "Top gaps": gbp_signal_label(r["signals_json"]),
        "GBP reviews": r["rating_votes"] if r["rating_votes"] is not None else "",
        "Neg reviews": r["neg_reviews"] if r["neg_reviews"] is not None else "",
        "Photos": r["total_photos"] if r["total_photos"] is not None else "",
        "Days since post": dsp if dsp is not None else "never",
        "Rating": r["rating"],
        "Phone": r["phone"] or "",
        "Email": r["email"] or "",
        "Owner": r["owner_name"] or "",
        "Website": r["website"] or "",
        "Google Maps": r["maps_url"] or "",
        "State": (r["state_code"] or "Unknown"),
        "place_id": r["place_id"],
    }


def to_record(r, min_reviews=DEFAULT_MIN_REVIEWS):
    """One ranked lead as a display record keyed by column name (shared by xlsx + CSV)."""
    _, qualified, signal = budget_assess(r, min_reviews)
    return {
        "Qualified": qualified,
        "Budget signal": signal,
        "Tier": _TIER.get(r["relevance"], ("?", 9))[0],
        "Business Name": r["name"],
        "Reviews": r["review_count"] or 0,
        "Rating": r["rating"],
        "Has Website": "NO" if r["no_website"] else "yes",
        "Site status": r["site_status"] or ("none" if r["no_website"] else ""),
        "Readiness": _readiness_cell(r["no_website"], r["site_status"], r["readiness_score"]),
        "Management": r["mgmt_status"] or "",
        "Conf": r["mgmt_confidence"] or "",
        "Builder": r["builder"] or "",
        "Marketing tags": _tags(r["marketing_tags_json"]),
        "Agency credit": r["agency_credit"] or "",
        "Top gap": _top_gap(r["no_website"], r["seo_gaps_json"]),
        "Phone": r["phone"] or "",
        "Email": r["email"] or "",
        "Owner": r["owner_name"] or "",
        "Website": r["website"] or "",
        "Google Maps": r["maps_url"] or "",
        "State": (r["state_code"] or "Unknown"),
        "place_id": r["place_id"],
    }


def group_by_state(rows, min_reviews=DEFAULT_MIN_REVIEWS, record_fn=to_record):
    """Ranked rows -> {state_code: [records]}, states ordered by lead count desc then
    code. Insertion order of each list preserves the within-state ranking. `record_fn`
    selects the per-lens record shape (to_record / to_gbp_record)."""
    groups = {}
    for r in rows:
        rec = record_fn(r, min_reviews)
        groups.setdefault(rec["State"], []).append(rec)
    ordered = sorted(groups.items(), key=lambda kv: (-len(kv[1]), kv[0]))
    return dict(ordered)


# ---- output writers ----

_HEAD_FILL = PatternFill("solid", fgColor="1F3864")


def _write_lead_sheet(ws, records, columns=LEAD_COLUMNS):
    ws.append(["#"] + columns)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = _HEAD_FILL
        c.alignment = Alignment(vertical="center")
    maps_idx = columns.index("Google Maps") if "Google Maps" in columns else None
    for i, rec in enumerate(records, start=1):
        row = [i] + [rec.get(col, "") for col in columns]
        if maps_idx is not None and rec.get("Google Maps"):
            row[1 + maps_idx] = f'=HYPERLINK("{rec["Google Maps"]}","map")'
        ws.append(row)
    for name, w in _WIDTHS.items():
        if name in columns:
            ws.column_dimensions[get_column_letter(columns.index(name) + 2)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns) + 1)}{len(records) + 1}"


def _gbp_summary_lines(groups):
    """Summary for the GBP-neglect lens (records carry GBP columns, not enrichment)."""
    total = sum(len(v) for v in groups.values())
    recs = [x for v in groups.values() for x in v]
    unclaimed = sum(1 for x in recs if x["Claimed"] == "NO")
    scored = [x["Neglect"] for x in recs if isinstance(x["Neglect"], (int, float))]
    avg = round(sum(scored) / len(scored), 1) if scored else 0
    lines = [
        (
            "GBP neglect audit - prospects ranked by profile neglect (worst first = best pitch)",
            None,
        ),
        ("", None),
        (f"{total} audited leads   |   {unclaimed} UNCLAIMED   |   avg neglect {avg}", None),
        ("", None),
        (f"By state/province ({len(groups)} tab{'s' if len(groups) != 1 else ''}):", None),
    ]
    for st, r in groups.items():
        nu = sum(1 for x in r if x["Claimed"] == "NO")
        lines.append((f"   {st}", f"{len(r)} leads ({nu} unclaimed)"))
    return lines


def _summary_lines(conn, groups, lens="opportunity", min_reviews=DEFAULT_MIN_REVIEWS):
    if lens == "gbp":
        return _gbp_summary_lines(groups)
    enr = leads_db.enrichment_status(conn)
    rel = {
        r[0]: r[1]
        for r in conn.execute("SELECT relevance, COUNT(*) FROM businesses GROUP BY relevance")
    }
    total = sum(len(v) for v in groups.values())
    no_site = sum(1 for recs in groups.values() for x in recs if x["Has Website"] == "NO")
    title = (
        "Budget-qualified prospects"
        if lens == "budget"
        else "Prospect report - all leads, ranked by opportunity"
    )
    lines = [
        (title, None),
        ("", None),
        (
            f"{total} leads   |   {no_site} with NO website   |   {total - no_site} with a website",
            None,
        ),
        (
            f"Relevance: {rel.get('match', 0)} CONFIRMED (trade-verified)   |   "
            f"{rel.get('maybe', 0)} REVIEW (general-contractor catch-all - eyeball before calling)",
            None,
        ),
    ]
    if lens == "budget":
        tiers = Counter(x["Qualified"] for recs in groups.values() for x in recs)
        lines += [
            ("", None),
            (f"Budget tiers (demand tier needs >= {min_reviews} reviews + a web gap):", None),
        ]
        for label in ("PROVEN BUDGET", "QUALIFIED DEMAND", "long tail"):
            if tiers.get(label):
                lines.append((f"   {label}", tiers[label]))
        lines.append(("", None))
        lines.append(
            ("Note: paid signals (ads/agency/stack) come from enrichment; un-enriched", None)
        )
        lines.append(
            ("leads can only qualify via no-website + reviews. Enrich for full coverage.", None)
        )
    lines += [
        ("", None),
        (f"By state/province ({len(groups)} tab{'s' if len(groups) != 1 else ''}):", None),
    ]
    for st, recs in groups.items():
        nw = sum(1 for x in recs if x["Has Website"] == "NO")
        lines.append((f"   {st}", f"{len(recs)} leads ({nw} no website)"))
    lines += [("", None), ("Website management mix (leads with a website):", None)]
    for k, v in enr["by_mgmt"].items():
        lines.append((f"   {k}", v))
    return lines


def write_workbook(
    conn,
    groups,
    out_path,
    columns=LEAD_COLUMNS,
    lens="opportunity",
    min_reviews=DEFAULT_MIN_REVIEWS,
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    for i, (text, val) in enumerate(_summary_lines(conn, groups, lens, min_reviews), start=1):
        ws.cell(row=i, column=1, value=text)
        if val is not None:
            ws.cell(row=i, column=2, value=val)
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 26

    for st, records in groups.items():
        _write_lead_sheet(wb.create_sheet(f"{st} ({len(records)})"[:31]), records, columns)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_csv(groups, csv_path, columns=LEAD_COLUMNS):
    """Flat master CSV of every lead (FR-14): State + within-state rank + lead columns
    + place_id, in the same ranked order as the tabs."""
    Path(csv_path).parent.mkdir(parents=True, exist_ok=True)
    header = ["State", "Rank"] + columns + ["place_id"]
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        for st, records in groups.items():
            for i, rec in enumerate(records, start=1):
                w.writerow([st, i] + [rec.get(c, "") for c in columns] + [rec["place_id"]])


def parse_region(arg):
    s = (arg or "").strip().lower()
    if not s or s == "all":
        return None
    return {c.strip() for c in s.replace(",", " ").split() if c.strip()}


def run(args):
    conn = leads_db.connect(args.db)
    leads_db.init_db(conn)
    region = parse_region(args.region)

    if args.lens == "gbp":
        rows = fetch_gbp_rows(conn, region)
        if not rows:
            print("No GBP-audited leads - run gbp_audit.py first (then re-run this report).")
            return 0
        columns = GBP_COLUMNS
        out = args.out or "output/leads-gbp-neglect.xlsx"
        groups = group_by_state(rows, args.min_reviews, record_fn=to_gbp_record)
        write_workbook(conn, groups, out, columns, args.lens, args.min_reviews)
        csv_path = args.csv or str(Path(out).with_suffix(".csv"))
        write_csv(groups, csv_path, columns)
        unclaimed = sum(1 for v in groups.values() for x in v if x["Claimed"] == "NO")
        print(
            f"[gbp lens] wrote {len(rows)} audited leads across {len(groups)} state tab(s); "
            f"{unclaimed} unclaimed"
        )
        print(f"  workbook -> {out}\n  master CSV -> {csv_path}")
        return 0

    rows = fetch_rows(conn, region)
    if not rows:
        print("No leads match - nothing to report.")
        return 0

    columns = LEAD_COLUMNS
    if args.lens == "budget":
        rows.sort(key=lambda r: _budget_sort_key(r, args.min_reviews))
        if args.qualified_only:
            rows = [r for r in rows if budget_assess(r, args.min_reviews)[0] <= 2]
            if not rows:
                print(
                    "No budget-qualified leads (no paid signals and none meet the review threshold)."
                )
                return 0
        columns = BUDGET_COLUMNS

    out = args.out or (
        "output/leads-budget-qualified.xlsx"
        if args.lens == "budget"
        else "output/leads-ranked.xlsx"
    )
    groups = group_by_state(rows, args.min_reviews)
    write_workbook(conn, groups, out, columns, args.lens, args.min_reviews)
    csv_path = args.csv or str(Path(out).with_suffix(".csv"))
    write_csv(groups, csv_path, columns)

    states = ", ".join(f"{st}={len(v)}" for st, v in groups.items())
    print(
        f"[{args.lens} lens] wrote {len(rows)} leads across {len(groups)} state tab(s) [{states}]"
    )
    if args.lens == "budget":
        tiers = Counter(x["Qualified"] for recs in groups.values() for x in recs)
        print(
            f"   PROVEN BUDGET: {tiers.get('PROVEN BUDGET', 0)} | "
            f"QUALIFIED DEMAND (>= {args.min_reviews} rev): {tiers.get('QUALIFIED DEMAND', 0)} | "
            f"long tail: {tiers.get('long tail', 0)}"
        )
    print(f"  workbook -> {out}")
    print(f"  master CSV -> {csv_path}")
    return 0


def main():
    ap = argparse.ArgumentParser(
        description="Segmented prospect report: per-state tabs + master CSV (FR-14)."
    )
    ap.add_argument("--db", default=leads_db.DEFAULT_DB)
    ap.add_argument(
        "--region", default="all", help="'all' or state/province codes, e.g. \"TN KY\"."
    )
    ap.add_argument(
        "--lens",
        choices=("opportunity", "budget", "gbp"),
        default="opportunity",
        help="opportunity = need-first (default); budget = who-can-pay first; "
        "gbp = most-neglected Google Business Profile first (needs gbp_audit.py).",
    )
    ap.add_argument(
        "--min-reviews",
        type=int,
        default=DEFAULT_MIN_REVIEWS,
        help="Min reviews for the budget lens 'qualified demand' tier (default 20).",
    )
    ap.add_argument(
        "--qualified-only",
        action="store_true",
        help="Budget lens: drop the long-tail tier, keep only proven-budget + qualified-demand.",
    )
    ap.add_argument("--out", default=None, help="Output .xlsx path (default depends on lens).")
    ap.add_argument(
        "--csv", default=None, help="Master CSV path (default: alongside --out as .csv)."
    )
    ap.add_argument(
        "--config", default=DEFAULT_CONFIG, help="Trade config (reserved for label use)."
    )
    args = ap.parse_args()
    return run(args)


if __name__ == "__main__":
    sys.exit(main())
