#!/usr/bin/env python3
"""
hvac_report.py — assemble a lead-gen / website-opportunity report for prospecting.

Repurposes the audit data for SALES, not benchmarking: it ranks each company by
a website-OPPORTUNITY score (weak site = strong prospect for web-upgrade
services) computed from measured signals, and merges in scraped contact data.
Deterministic tables only; the agent writes the narrative + per-lead pitch.

Inputs (all under .tmp/): discovery JSON (companies), audit/*.json, contacts.json.
Usage:
    python tools/hvac_report.py .tmp/discover/<slug>.json --out output/<slug>-prospects.md
"""

import argparse
import json
import sys
from pathlib import Path
from urllib.parse import urlparse

sys.path.insert(0, str(Path(__file__).resolve().parent))
from score_report import find_audit  # reuse domain matching

# Default industry schema to look for; overridden per-run from the discovery
# JSON's "industry_schema" (set in the targets config) so the tool is reusable
# across trades. "localbusiness" + the vertical's own subtype, if schema.org has one.
DEFAULT_LOCAL_SCHEMA = ("localbusiness", "hvacbusiness")


def homepage(audit):
    pages = [p for p in audit.get("pages", {}).values() if not p.get("error")]
    return pages[0] if pages else None


def opportunity(hp, site, local_schema=DEFAULT_LOCAL_SCHEMA, vertical="HVAC"):
    """Higher score = weaker website = better prospect. Returns (score, gaps[]).
    score=None means the site has a URL but could not be audited (timeout/block)."""
    if not hp:
        return None, ["site did not load — could not audit; recheck manually"]
    gaps, score = [], 0
    types = [t.lower() for t in hp.get("schema_types", [])]
    if not hp.get("https"):
        score += 3
        gaps.append("no HTTPS")
    if not hp.get("viewport"):
        score += 3
        gaps.append("not mobile-friendly (no viewport)")
    if not types:
        score += 3
        gaps.append("no structured data (JSON-LD) — weak for Google & AI search")
    else:
        if not any(s in t for t in types for s in local_schema):
            score += 1
            gaps.append(f"no LocalBusiness/{vertical} schema")
        if "faqpage" not in types:
            score += 1
            gaps.append("no FAQ schema")
    tl = hp.get("title_length", 0)
    if not (10 <= tl <= 65):
        score += 1
        gaps.append(f"title length off ({tl}c)")
    dl = hp.get("meta_description_length", 0)
    if not (50 <= dl <= 170):
        score += 1
        gaps.append(f"meta description off ({dl}c)")
    w = hp.get("word_count", 0)
    if w < 400:
        score += 2
        gaps.append(f"thin homepage content ({w} words)")
    elif w < 600:
        score += 1
        gaps.append(f"light homepage content ({w} words)")
    if not hp.get("canonical"):
        score += 1
        gaps.append("no canonical tag")
    return score, gaps


SOCIAL_HOSTS = (
    "facebook.com",
    "instagram.com",
    "m.facebook.com",
    "linktr.ee",
    "linktree",
    "business.site",
)
DIRECTORY_HOSTS = (
    "bbb.org",
    "yelp.com",
    "yellowpages.com",
    "angi.com",
    "mapquest.com",
    "manta.com",
    "porch.com",
    "thumbtack.com",
    "nextdoor.com",
)
PARKED_HINTS = (
    "domain for sale",
    "is for sale",
    "buy this domain",
    "parked",
    "godaddy.com",
    "hugedomains",
    "site not found",
    "website coming soon",
    "default web page",
)


def classify_site(host, hp, no_website):
    """Return (status, note). status in: none, live, dead, social_only, directory,
    parked, blocked, unreachable. Only 'live' gets an opportunity score; dead/social/
    directory/parked count as 'no usable website' (build leads); blocked = couldn't audit."""
    if no_website or not host:
        return "none", "no website listed"
    h = host.lower()
    if any(s in h for s in SOCIAL_HOSTS):
        return "social_only", "social page only (no real website)"
    if any(d in h for d in DIRECTORY_HOSTS):
        return "directory", "directory listing only (no real website)"
    if hp is None or hp.get("error"):
        return "unreachable", "domain unreachable (DNS/connection) — likely dead, verify"
    st = hp.get("status")
    title = (hp.get("title") or "").lower()
    wc = hp.get("word_count", 0)
    if any(k in title for k in PARKED_HINTS):
        return "parked", f'parked/placeholder page ("{(hp.get("title") or "")[:40]}")'
    if st in (404, 410) or (("not found" in title or "404" in title) and wc < 120):
        return "dead", f'dead page (HTTP {st}: "{(hp.get("title") or "")[:35]}")'
    if st in (200, 201):
        return "live", ""
    if st in (403, 429, 202, 503) or (isinstance(st, int) and 500 <= st < 600) or st == 400:
        return "blocked", f"live but couldn't audit (HTTP {st} — WAF/JS block); verify by hand"
    return "blocked", f"unexpected HTTP {st}; verify by hand"


def site_signals(audit):
    s = audit.get("site", {})
    return {
        "sitemap": bool(
            s.get("robots_txt", {}).get("sitemap_declared")
            or s.get("sitemap_xml", {}).get("exists")
        ),
        "llms": bool(s.get("llms_txt", {}).get("exists")),
        "bots_blocked": [
            b for b, v in s.get("robots_txt", {}).get("ai_bots", {}).items() if v == "blocked"
        ],
    }


VALID_TLDS = {
    "com",
    "net",
    "org",
    "co",
    "us",
    "biz",
    "info",
    "io",
    "us.com",
    "email",
    "pro",
    "company",
    "services",
    "tech",
}


def best_email(rec):
    if not rec:
        return None
    junk = {
        "your@email.com",
        "email@example.com",
        "info@example.com",
        "john@gmail.com",
        "test@test.com",
    }

    def valid(e):
        if e in junk or e.count("@") != 1:
            return False
        local, domain = e.split("@")
        if "." not in domain:
            return False
        # reject reversed/obfuscated anti-scrape strings (e.g. "moc.x@y.obf" = .com reversed)
        if local.startswith(("moc.", "ten.", "gro.")) or domain.split(".")[-1] not in VALID_TLDS:
            return False
        return True

    cands = [e for e in rec.get("emails", []) if valid(e)]
    return cands[0] if cands else None


COLUMNS = [
    ("group", "Group"),
    ("rank", "Rank"),
    ("name", "Company"),
    ("tier", "Tier"),
    ("score", "Opportunity"),
    ("rating", "Google rating"),
    ("reviews", "Reviews"),
    ("phone", "Phone"),
    ("email", "Email"),
    ("owner", "Owner (if public)"),
    ("website", "Website"),
    ("site_note", "Site status"),
    ("address", "Area / address"),
    ("mobile", "Mobile-friendly"),
    ("https", "HTTPS"),
    ("words", "Homepage words"),
    ("gaps_str", "Top gaps"),
    ("schema", "Schema types"),
    ("found_via", "Found via (towns)"),
    ("maps_url", "Google Maps"),
    ("notes", "Notes"),
]


def _flat(r, group, rank):
    def cell(v):
        if v is None:
            return ""
        if isinstance(v, bool):
            return "yes" if v else "no"
        return v

    out = dict(r)
    out["group"] = group
    out["rank"] = rank
    out["owner"] = r.get("owner") or "not found"
    out["email"] = r.get("email") or "not found"
    out["phone"] = r.get("phone") or "not found"
    out["site_note"] = r.get("site_note") or ("live" if r.get("site_status") == "live" else "")
    out["found_via"] = (
        ", ".join(r["found_via"])
        if isinstance(r.get("found_via"), list)
        else (r.get("found_via") or "")
    )
    if r.get("no_website"):
        out["score"] = "NO SITE"
        out["gaps_str"] = "no website — full build opportunity"
    elif r.get("score") is None:
        out["score"] = "n/a (site error)"
        out["gaps_str"] = "; ".join(r.get("gaps", []))
    else:
        out["gaps_str"] = "; ".join(r.get("gaps", []))
    return {key: cell(out.get(key, "")) for key, _ in COLUMNS}


def write_csv(path, groups):
    import csv

    rows = [_flat(r, g, i) for g, items in groups for i, r in enumerate(items, 1)]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:  # utf-8-sig so Excel shows accents
        w = csv.DictWriter(f, fieldnames=[k for k, _ in COLUMNS])
        w.writerow({k: label for k, label in COLUMNS})
        for row in rows:
            w.writerow(row)


def write_xlsx(path, groups):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(
            "  (openpyxl not installed — skipping .xlsx; CSV written. `pip install openpyxl` for Excel.)"
        )
        return False
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    labels = [label for _, label in COLUMNS][1:]  # drop 'Group' col inside per-sheet tabs
    keys = [k for k, _ in COLUMNS][1:]
    for sheet_name, items in groups:
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.append(labels)
        for c in range(1, len(labels) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        for i, r in enumerate(items, 1):
            flat = _flat(r, sheet_name, i)
            ws.append([flat[k] for k in keys])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(labels))}{ws.max_row}"
        # column widths
        widths = {
            "Company": 34,
            "Email": 30,
            "Owner (if public)": 26,
            "Website": 38,
            "Area / address": 30,
            "Top gaps": 60,
            "Schema types": 44,
            "Found via (towns)": 30,
            "Notes": 40,
        }
        for c, label in enumerate(labels, 1):
            ws.column_dimensions[get_column_letter(c)].width = widths.get(label, 14)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    try:
        wb.save(path)
        return path
    except PermissionError:
        alt = Path(path).with_name(Path(path).stem + "-new.xlsx")
        wb.save(alt)
        print(
            f"  (note: {Path(path).name} was locked/open in Excel — wrote {alt.name} instead. Close Excel and rename if you like.)"
        )
        return alt


def main():
    ap = argparse.ArgumentParser(
        description="Assemble a website-opportunity prospect report (vertical-agnostic; label/schema from the discovery JSON)."
    )
    ap.add_argument("discover")
    ap.add_argument("--audit-dir", default=".tmp/audit")
    ap.add_argument("--contacts", default=".tmp/contacts/contacts.json")
    ap.add_argument("--out", default=None)
    ap.add_argument(
        "--no-md",
        action="store_true",
        help="Skip the markdown (e.g. to refresh CSV/XLSX without clobbering a hand-edited narrative).",
    )
    args = ap.parse_args()

    data = json.loads(Path(args.discover).read_text(encoding="utf-8"))
    contacts = (
        json.loads(Path(args.contacts).read_text(encoding="utf-8"))
        if Path(args.contacts).exists()
        else {}
    ).get("contacts", {})

    # Vertical-driven labels/schema (defaults keep prior HVAC behaviour intact).
    vertical = data.get("vertical") or "HVAC"
    local_schema = tuple(s.lower() for s in (data.get("industry_schema") or DEFAULT_LOCAL_SCHEMA))

    rows = []
    for c in data.get("companies", []):
        listed_no_web = c.get("no_website") or not c.get("website")
        host = (
            urlparse(c["website"]).netloc if c.get("website") else ""
        )  # clean host; ignores ?utm_ params
        audit = None if listed_no_web else find_audit(host, args.audit_dir)
        hp = homepage(audit) if audit else None
        sig = (
            site_signals(audit) if audit else {"sitemap": False, "llms": False, "bots_blocked": []}
        )
        status, status_note = classify_site(host, hp, listed_no_web)
        # "no usable website" = no site at all, OR a site that's dead/social/directory/parked
        no_usable = status in ("none", "dead", "social_only", "directory", "parked", "unreachable")

        if status == "live":
            score, gaps = opportunity(hp, sig, local_schema, vertical)
            if score is not None:
                if not sig["sitemap"]:
                    score += 1
                    gaps.append("no sitemap")
                if not sig["llms"]:
                    score += 1
                    gaps.append("no llms.txt (AI-search file)")
        else:
            score, gaps = None, ([status_note] if status_note else [])

        cr = contacts.get(c["name"], {})
        rows.append(
            {
                "name": c["name"],
                "tier": c.get("tier", "?"),
                "website": c.get("website"),
                "no_website": no_usable,
                "site_status": status,
                "site_note": status_note,
                "phone": c.get("phone"),
                "email": c.get("email") or best_email(cr),
                "owner": c.get(
                    "owner"
                ),  # only set when publicly verified; else None -> "not found"
                "address": c.get("address"),
                "rating": c.get("rating"),
                "reviews": c.get("review_count"),
                "score": score,
                "gaps": gaps,
                "words": hp.get("word_count") if hp else None,
                "schema": ", ".join(hp.get("schema_types", []))
                if (hp and status == "live")
                else "",
                "https": hp.get("https") if (hp and status == "live") else None,
                "mobile": hp.get("viewport") if (hp and status == "live") else None,
                "found_via": c.get("found_via", []),
                "maps_url": c.get("maps_url"),
                "notes": c.get("notes", ""),
            }
        )

    def rev(r):
        return r.get("reviews") or 0

    build_leads = sorted([r for r in rows if r["no_website"]], key=rev, reverse=True)
    local = sorted(
        [r for r in rows if not r["no_website"] and r["tier"] == "local"],
        key=lambda r: (r["score"] if r["score"] is not None else -1, rev(r)),
        reverse=True,
    )
    chains = sorted(
        [r for r in rows if not r["no_website"] and r["tier"] == "chain"], key=rev, reverse=True
    )
    none_n = sum(1 for r in build_leads if r["site_status"] == "none")
    broken_n = len(build_leads) - none_n
    blocked_n = sum(1 for r in rows if r["site_status"] == "blocked")

    L = []
    L.append(f"# {vertical} Website-Opportunity & Contact Report — {data.get('area', '')}\n")
    L.append(
        f"> **Purpose:** prospect list for selling website-upgrade/creation services — every {vertical} company found in the target area, with verified Google ratings, contact details, and (for those with a working site) a website-weakness ranking. Weakest/absent web presence = best pitch.\n"
    )
    L.append(
        f"*{len(rows)} businesses · {len(build_leads)} need a website built ({none_n} have none, {broken_n} have a dead/social/placeholder link) · {len(local)} local with a working site · {len(chains)} regional/national.*\n"
    )

    L.append("## How to read this / limitations\n")
    L.append(f"- **Discovery method:** {data.get('source', 'Google Places API')}.")
    L.append(
        "- **Every listed website was re-checked with a real browser** (an earlier pass used a crawler User-Agent that many security firewalls blocked, making live sites look broken — now fixed). The **Site status** column records the result: *live*, *dead (HTTP 404)*, *social page only*, *directory listing only*, *parked*, or *blocked (couldn't audit — verify by hand)*."
    )
    L.append(
        '- **Build-from-scratch leads come first** — companies with no website **plus** companies whose only "website" is dead, a Facebook page, or a directory listing. Sorted by Google review count (proof they\'re real and active).'
    )
    L.append(
        "- **Website-opportunity score** (companies with a *working* site) sums measured weaknesses (no HTTPS/mobile/schema/FAQ, thin content, bad title/meta, no sitemap/llms.txt). Higher = weaker = bigger opportunity. Not a Google ranking."
    )
    L.append(
        "- **Ratings/reviews** are live Google Places data. **Contact data** is scraped from each public site; owner names/emails appear only where publicly posted, else *not found* (no guessing)."
    )
    L.append(
        "- **Regional/national operators** are tagged separately but fully audited — several are franchises with a local owner/manager, and a big name can still have a weak site."
    )
    L.append(
        f"- **{blocked_n} sites returned a firewall/JS block** even to a browser and are marked *blocked* with no score — verify those by hand."
    )
    L.append(
        "- **Not measured:** AI-search visibility (no Perplexity key) and backlinks/SERP (no DataForSEO key).\n"
    )

    def contact_table(title, items, show_opp=True):
        L.append(f"## {title}\n")
        if show_opp:
            L.append(
                "| # | Company | Opp | Rating | Reviews | Phone | Email | Owner (if public) | Website | Site status |"
            )
            L.append("|--:|---|--:|--:|--:|---|---|---|---|---|")
        else:
            L.append(
                "| # | Company | Rating | Reviews | Phone | Email | Owner (if public) | Listed link | Why it needs a site |"
            )
            L.append("|--:|---|--:|--:|---|---|---|---|---|")
        for i, r in enumerate(items, 1):
            owner = r.get("owner") or "*not found*"
            rating = f"{r['rating']}★" if r.get("rating") else "—"
            reviews = r.get("reviews") if r.get("reviews") is not None else "—"
            note = r.get("site_note") or ("live" if r.get("site_status") == "live" else "—")
            if show_opp:
                opp = "n/a" if r["score"] is None else f"**{r['score']}**"
                L.append(
                    f"| {i} | {r['name']} | {opp} | {rating} | {reviews} | {r['phone'] or '—'} | "
                    f"{r['email'] or '*nf*'} | {owner} | {r['website'] or '—'} | {note} |"
                )
            else:
                L.append(
                    f"| {i} | {r['name']} | {rating} | {reviews} | {r['phone'] or '—'} | "
                    f"{r['email'] or '*nf*'} | {owner} | {r['website'] or '— (none listed)'} | {note} |"
                )
        L.append("")

    contact_table(
        "⭐ Build-from-scratch leads (no usable website — highest opportunity)",
        build_leads,
        show_opp=False,
    )

    contact_table(
        "Local prospects with a working site (ranked by site weakness)", local, show_opp=True
    )
    L.append("### Top gaps for the weakest local sites\n")
    for r in [x for x in local if x["score"] is not None][:15]:
        L.append(
            f"- **{r['name']}** (opp {r['score']}): {'; '.join(r['gaps'][:4]) or 'solid site'}"
        )
    L.append("")

    if chains:
        contact_table(
            "Regional / national operators (larger — fully audited too)", chains, show_opp=True
        )
        L.append("### Top gaps for the weakest chain/regional sites\n")
        for r in [x for x in chains if x["score"] is not None][:12]:
            L.append(
                f"- **{r['name']}** (opp {r['score']}): {'; '.join(r['gaps'][:4]) or 'solid site'}"
            )
        L.append("")

    L.append("## Agent narrative\n")
    L.append("> _Filled in below by the agent: per-lead pitch angle grounded in the gaps above._\n")

    out = (
        Path(args.out) if args.out else Path("output") / f"{data.get('slug', 'hvac')}-prospects.md"
    )
    out.parent.mkdir(parents=True, exist_ok=True)
    if not args.no_md:
        out.write_text("\n".join(L), encoding="utf-8")
        print(f"Wrote {out}")

    groups = [
        ("Build-from-scratch leads", build_leads),
        ("Local w- working site", local),
        ("Chains", chains),
    ]
    csv_path = out.with_suffix(".csv")
    write_csv(csv_path, groups)
    print(f"Wrote {csv_path}")
    written = write_xlsx(out.with_suffix(".xlsx"), groups)
    if written:
        print(f"Wrote {written}")
    print(
        f"  {len(build_leads)} build-from-scratch leads · {len(local)} local w/ working site · {len(chains)} chains"
    )


if __name__ == "__main__":
    main()
