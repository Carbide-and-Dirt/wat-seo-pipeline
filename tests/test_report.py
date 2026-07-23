#!/usr/bin/env python3
"""
Tests for leads_report.py - the ranked prospect workbook export (HLD FR-14/18).

Pure logic on a temp SQLite DB (no network, no spend, no xlsx written): guards the
opportunity RANKING, which is the part that silently produces a misleading call list
if it regresses. Run standalone or under pytest:

    python tests/test_report.py
    pytest tests/
"""

import json
import sys
import tempfile
from pathlib import Path

TOOLS = Path(__file__).resolve().parent.parent / "tools"
sys.path.insert(0, str(TOOLS))

import leads_db
import leads_report as lr


def _db():
    conn = leads_db.connect(str(Path(tempfile.mkdtemp()) / "t.sqlite"))
    leads_db.init_db(conn)
    return conn


def _biz(conn, pid, name, relevance, no_website, reviews):
    conn.execute(
        "INSERT INTO businesses (place_id, name, website, no_website, state_code, "
        "review_count, relevance) VALUES (?,?,?,?,?,?,?)",
        (
            pid,
            name,
            "" if no_website else f"https://{pid}.com",
            no_website,
            "TN",
            reviews,
            relevance,
        ),
    )
    conn.commit()


def _enrich(conn, pid, site_status, readiness=None):
    leads_db.upsert_enrichment(
        conn,
        {"place_id": pid, "site_status": site_status, "readiness_score": readiness},
        "2026-06-18T00:00:00+00:00",
    )


def test_opp_bucket_order():
    # no-site < broken < live < blocked
    assert lr._opp_bucket(1, None) == 0
    assert lr._opp_bucket(0, "dead") == 1
    assert lr._opp_bucket(0, "social_only") == 1
    assert lr._opp_bucket(0, "live") == 2
    assert lr._opp_bucket(0, "blocked") == 4


def test_full_ranking_order():
    conn = _db()
    # CONFIRMED tier
    _biz(conn, "c_nw_lo", "C no-site few reviews", "match", 1, 5)
    _biz(conn, "c_nw_hi", "C no-site many reviews", "match", 1, 90)
    _biz(conn, "c_live_weak", "C live weak", "match", 0, 3)
    _enrich(conn, "c_live_weak", "live", 12)
    _biz(conn, "c_live_strong", "C live strong", "match", 0, 50)
    _enrich(conn, "c_live_strong", "live", 1)
    _biz(conn, "c_dead", "C dead site", "match", 0, 8)
    _enrich(conn, "c_dead", "dead")
    _biz(conn, "c_blocked", "C blocked", "match", 0, 200)
    _enrich(conn, "c_blocked", "blocked")
    # REVIEW tier (a no-website lead, so no enrichment row - LEFT JOIN yields NULLs).
    # Must sort AFTER every CONFIRMED row despite having the most reviews of all.
    _biz(conn, "r_nw", "R no-site huge reviews", "maybe", 1, 999)

    order = [r["name"] for r in lr.fetch_rows(conn)]
    assert order == [
        "C no-site many reviews",  # tier match, bucket 0, reviews desc
        "C no-site few reviews",
        "C dead site",  # bucket 1 (effectively no site)
        "C live weak",  # bucket 2, readiness 12 (weaker first)
        "C live strong",  # bucket 2, readiness 1
        "C blocked",  # bucket 4, last within tier despite 200 reviews
        "R no-site huge reviews",  # whole REVIEW tier sorts last, even with 999 reviews
    ], order


def test_region_filter():
    conn = _db()
    _biz(conn, "tn1", "TN biz", "match", 1, 1)
    conn.execute("UPDATE businesses SET state_code='KY' WHERE place_id='tn1'")
    conn.commit()
    assert lr.fetch_rows(conn, state_codes={"tn"}) == []
    assert len(lr.fetch_rows(conn, state_codes={"ky"})) == 1


# ---- FR-14: per-state grouping + CSV + workbook tabs ----


def _multi_state_db():
    conn = _db()
    _biz(conn, "tn_hi", "TN busy", "match", 1, 80)
    _biz(conn, "tn_lo", "TN quiet", "match", 1, 5)
    _biz(conn, "ky_one", "KY one", "match", 1, 50)
    conn.execute(
        "UPDATE businesses SET state_code='KY', state_name='Kentucky' WHERE place_id='ky_one'"
    )
    conn.execute("UPDATE businesses SET state_name='Tennessee' WHERE state_code='TN'")
    conn.commit()
    return conn


def test_group_by_state_orders_by_count_then_preserves_rank():
    conn = _multi_state_db()
    groups = lr.group_by_state(lr.fetch_rows(conn))
    assert list(groups.keys()) == ["TN", "KY"]  # TN (2) before KY (1)
    assert [r["Business Name"] for r in groups["TN"]] == ["TN busy", "TN quiet"]  # reviews desc


def test_to_record_no_website_fields():
    conn = _db()
    _biz(conn, "p1", "No Site Co", "match", 1, 9)
    rec = lr.to_record(lr.fetch_rows(conn)[0])
    assert rec["Tier"] == "CONFIRMED" and rec["Has Website"] == "NO"
    assert rec["Readiness"] == "NO SITE" and rec["State"] == "TN"
    assert rec["Top gap"] == "no website - full build opportunity"


def test_write_csv_has_state_rank_and_all_rows():
    conn = _multi_state_db()
    groups = lr.group_by_state(lr.fetch_rows(conn))
    out = Path(tempfile.mkdtemp()) / "r.csv"
    lr.write_csv(groups, str(out))
    lines = out.read_text(encoding="utf-8").splitlines()
    assert lines[0].startswith("State,Rank,Tier,Business Name")
    assert len(lines) == 1 + 3  # header + 3 leads
    assert lines[1].startswith("TN,1,")  # first TN lead is rank 1


def test_write_workbook_one_tab_per_state():
    import openpyxl

    conn = _multi_state_db()
    groups = lr.group_by_state(lr.fetch_rows(conn))
    out = Path(tempfile.mkdtemp()) / "r.xlsx"
    lr.write_workbook(conn, groups, str(out))
    wb = openpyxl.load_workbook(out, read_only=True)
    assert wb.sheetnames == ["Summary", "TN (2)", "KY (1)"]
    ws = wb["TN (2)"]
    assert [c.value for c in ws[1]][:3] == ["#", "Tier", "Business Name"]


# ---- budget-qualified lens ----


def _row(**kw):
    base = {
        "place_id": "pid",
        "google_ads": 0,
        "mgmt_status": None,
        "marketing_tags_json": None,
        "no_website": 0,
        "site_status": "live",
        "readiness_score": 0,
        "review_count": 0,
        "name": "X",
        "relevance": "match",
        "rating": None,
        "website": "https://x.com",
        "phone": None,
        "state_code": "TN",
        "state_name": "Tennessee",
        "maps_url": None,
        "seo_gaps_json": None,
        "mgmt_confidence": None,
        "builder": None,
        "agency_credit": None,
        "email": None,
        "owner_name": None,
    }
    base.update(kw)
    return base


def test_paid_signal_priority():
    assert lr.paid_signal(_row(google_ads=1, mgmt_status="DIY / unmanaged")) == "Google Ads"
    assert lr.paid_signal(_row(mgmt_status="likely agency-managed")) == "agency-managed"
    assert (
        lr.paid_signal(_row(marketing_tags_json=json.dumps(["CallRail", "Meta Pixel"])))
        == "CallRail"
    )
    assert (
        lr.paid_signal(_row(marketing_tags_json=json.dumps(["Meta Pixel"]))) == ""
    )  # not a managed-stack tag


def test_budget_assess_tiers():
    assert lr.budget_assess(_row(google_ads=1))[0] == 1  # proven budget
    t, label, why = lr.budget_assess(_row(no_website=1, review_count=40, site_status=None))
    assert t == 2 and "QUALIFIED" in label and "40 reviews" in why  # busy + no site
    assert (
        lr.budget_assess(_row(review_count=25, readiness_score=8))[0] == 2
    )  # busy + weak live site
    assert (
        lr.budget_assess(_row(review_count=99, readiness_score=0))[0] == 3
    )  # busy but solid site, no gap
    assert (
        lr.budget_assess(_row(no_website=1, review_count=3, site_status=None))[0] == 3
    )  # has gap but too small


def test_budget_sort_orders_tiers_and_within():
    rows = [
        _row(name="solid-busy", review_count=99, readiness_score=0),  # tier 3
        _row(name="ads-strongsite", google_ads=1, readiness_score=1),  # tier 1
        _row(name="ads-weaksite", google_ads=1, readiness_score=12),  # tier 1, weaker -> first
        _row(name="busy-nosite", no_website=1, review_count=50, site_status=None),  # tier 2
    ]
    rows.sort(key=lambda r: lr._budget_sort_key(r, 20))
    assert [r["name"] for r in rows] == [
        "ads-weaksite",
        "ads-strongsite",
        "busy-nosite",
        "solid-busy",
    ]


def test_to_record_carries_budget_fields():
    rec = lr.to_record(_row(google_ads=1, name="Acme"))
    assert rec["Qualified"] == "PROVEN BUDGET" and rec["Budget signal"] == "Google Ads"


# ---- GBP-neglect lens ----


def _gbp(conn, pid, neglect, signals, is_claimed=0, votes=None, photos=None, audit_type="prospect"):
    import leads_db_gbp as gdb

    gdb.insert_gbp_audit(
        conn,
        place_id=pid,
        audit_type=audit_type,
        fields={"is_claimed": is_claimed, "rating_votes": votes, "total_photos": photos},
        neglect_score=neglect,
        signals=signals,
        api_cost_usd=0.003,
        status="complete",
    )


def test_gbp_signal_label_orders_by_weight_and_truncates():
    lbl = lr.gbp_signal_label('{"few_photos": true, "unclaimed": true, "no_hours": true}', top=2)
    assert lbl == "unclaimed, few photos"  # unclaimed leads, capped at 2


def test_gbp_lens_ranks_most_neglected_first_latest_audit():
    conn = _db()
    _biz(conn, "A", "Low neglect now", "match", 0, 30)
    _biz(conn, "B", "High neglect", "match", 0, 10)
    _biz(conn, "C", "Maybe tier worst", "maybe", 0, 50)
    conn.execute("UPDATE businesses SET lat=36.1, lng=-86.7")  # gbp lens needs coordinates upstream
    _gbp(conn, "A", 90.0, {"unclaimed": True})  # older
    _gbp(conn, "A", 15.0, {"few_photos": True}, is_claimed=1, audit_type="monthly")  # newer wins
    _gbp(conn, "B", 85.0, {"unclaimed": True, "few_photos": True})
    _gbp(conn, "C", 95.0, {"unclaimed": True})
    conn.commit()
    order = [(r["name"], r["neglect_score"]) for r in lr.fetch_gbp_rows(conn)]
    # match tier first (B 85 before A's latest 15), then maybe tier (C 95)
    assert order == [("High neglect", 85.0), ("Low neglect now", 15.0), ("Maybe tier worst", 95.0)]


def test_to_gbp_record_and_summary():
    conn = _db()
    _biz(conn, "B", "Unclaimed biz", "match", 0, 10)
    _gbp(conn, "B", 85.0, {"unclaimed": True, "few_photos": True}, is_claimed=0, votes=4, photos=1)
    conn.commit()
    rec = lr.to_gbp_record(lr.fetch_gbp_rows(conn)[0])
    assert rec["Claimed"] == "NO" and rec["Neglect"] == 85 and "unclaimed" in rec["Top gaps"]
    groups = lr.group_by_state(lr.fetch_gbp_rows(conn), record_fn=lr.to_gbp_record)
    assert any("UNCLAIMED" in str(t) for t, _ in lr._gbp_summary_lines(groups))


def _run_all():
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    passed = 0
    for t in tests:
        try:
            t()
            print(f"  PASS  {t.__name__}")
            passed += 1
        except Exception as e:  # noqa: BLE001
            print(f"  FAIL  {t.__name__}: {type(e).__name__}: {e}")
    print(f"\n{passed}/{len(tests)} passed")
    return 0 if passed == len(tests) else 1


if __name__ == "__main__":
    sys.exit(_run_all())
